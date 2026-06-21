import os
import json
import uuid
from datetime import datetime, timedelta
from decimal import Decimal
from functools import wraps
from io import BytesIO

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from flask_jwt_extended import JWTManager, create_access_token, jwt_required, get_jwt_identity
from flask_socketio import SocketIO, emit, join_room, leave_room
import redis
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import logging

app = Flask(__name__)
CORS(app, resources={r"/api/*": {"origins": "*"}})

app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv(
    'DATABASE_URL',
    'postgresql://spp_user:spp_password@localhost:5432/spp_db'
)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

app.config['JWT_SECRET_KEY'] = os.getenv('JWT_SECRET_KEY', 'your-secret-key-change-in-production')
app.config['JWT_ACCESS_TOKEN_EXPIRES'] = timedelta(days=30)

REDIS_URL = os.getenv('REDIS_URL', 'redis://localhost:6379')
redis_client = redis.from_url(REDIS_URL, decode_responses=True)

db = SQLAlchemy(app)
jwt = JWTManager(app)
socketio = SocketIO(app, cors_allowed_origins="*", async_mode='eventlet')

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ============================================================================
# Database Models
# ============================================================================

class SPPElement(db.Model):
    __tablename__ = 'spp_elements'
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(50), unique=True, nullable=False)
    name = db.Column(db.String(255), nullable=False)
    description = db.Column(db.Text)
    parent_id = db.Column(db.Integer, db.ForeignKey('spp_elements.id'))
    status = db.Column(db.String(20), default='ACTIVE')
    level = db.Column(db.Integer, default=1)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow)

class Department(db.Model):
    __tablename__ = 'departments'
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(50), unique=True, nullable=False)
    name = db.Column(db.String(255), nullable=False)
    description = db.Column(db.Text)
    status = db.Column(db.String(20), default='ACTIVE')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class DistributionResult(db.Model):
    __tablename__ = 'distribution_results'
    id = db.Column(db.Integer, primary_key=True)
    session_id = db.Column(db.String(255), nullable=False)
    version_date = db.Column(db.DateTime, nullable=False)
    spp_version_id = db.Column(db.Integer)
    total_amount = db.Column(db.Numeric(15, 2), nullable=False)
    distribution_data = db.Column(db.JSON, nullable=False)
    status = db.Column(db.String(20), default='SAVED')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    exported_at = db.Column(db.DateTime)
    meta_info = db.Column('metadata', db.JSON)

class UserSession(db.Model):
    __tablename__ = 'user_sessions'
    id = db.Column(db.Integer, primary_key=True)
    session_id = db.Column(db.String(255), unique=True, nullable=False)
    user_id = db.Column(db.String(255))
    jwt_token = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    last_activity = db.Column(db.DateTime, default=datetime.utcnow)
    expires_at = db.Column(db.DateTime)

# ============================================================================
# Helper functions (all use db.text with .mappings() for dictionary access)
# ============================================================================

def round_to_two_decimals(value):
    return round(float(value), 2)

def get_element_info(element_id, version_date):
    """Возвращает словарь с информацией об элементе на указанную дату (SCD2)."""
    if version_date:
        query = db.text("""
            SELECT element_id as id, code, name, parent_id, level, status
            FROM spp_history
            WHERE element_id = :element_id
              AND valid_from <= :target_date
              AND (valid_to IS NULL OR valid_to > :target_date)
            LIMIT 1
        """)
        row = db.session.execute(query, {
            'element_id': element_id,
            'target_date': version_date
        }).mappings().fetchone()
        if row:
            return dict(row)   # словарь
    elem = SPPElement.query.get(element_id)
    if elem:
        return {
            'id': elem.id,
            'code': elem.code,
            'name': elem.name,
            'parent_id': elem.parent_id,
            'level': elem.level,
            'status': elem.status
        }
    return None

def get_ancestors(element_id, version_date):
    """Возвращает список id всех предков элемента на дату (SCD2)."""
    ancestors = []
    current_id = element_id
    while current_id is not None:
        parent_query = db.text("""
            SELECT h.parent_id
            FROM spp_history h
            WHERE h.element_id = :element_id
              AND h.valid_from <= :target_date
              AND (h.valid_to IS NULL OR h.valid_to > :target_date)
            LIMIT 1
        """)
        row = db.session.execute(parent_query, {
            'element_id': current_id,
            'target_date': version_date
        }).mappings().fetchone()
        parent_id = row['parent_id'] if row else None
        if parent_id is not None:
            ancestors.append(parent_id)
            current_id = parent_id
        else:
            break
    return ancestors

def filter_independent_elements(element_ids, version_date):
    """Оставляет только те элементы, у которых ни один предок не входит в element_ids."""
    independent = []
    for eid in element_ids:
        ancestors = get_ancestors(eid, version_date)
        if not any(anc in element_ids for anc in ancestors):
            independent.append(eid)
    return independent

def _collect_descendants(element_id, version_date, all_ids):
    """Рекурсивно добавляет id всех активных потомков в множество all_ids."""
    query = db.text("""
        SELECT element_id
        FROM spp_history
        WHERE parent_id = :parent_id
          AND valid_from <= :target_date
          AND (valid_to IS NULL OR valid_to > :target_date)
          AND status = 'ACTIVE'
    """)
    rows = db.session.execute(query, {
        'parent_id': element_id,
        'target_date': version_date
    }).mappings().fetchall()
    for row in rows:
        child_id = row['element_id']
        if child_id not in all_ids:
            all_ids.add(child_id)
            _collect_descendants(child_id, version_date, all_ids)

def build_full_tree(top_element_ids, version_date):
    """
    Строит полное дерево (словарь) от корней, содержащее всех предков
    и потомков для указанных top_element_ids.
    """
    all_ids = set(top_element_ids)
    for eid in top_element_ids:
        ancestors = get_ancestors(eid, version_date)
        all_ids.update(ancestors)
        _collect_descendants(eid, version_date, all_ids)

    nodes_info = {}
    for eid in all_ids:
        info = get_element_info(eid, version_date)
        if info:
            nodes_info[eid] = info

    # Определим корни (элементы, parent_id которых отсутствует в all_ids)
    roots = []
    children_map = {eid: [] for eid in all_ids}
    for eid, info in nodes_info.items():
        parent_id = info.get('parent_id')
        if parent_id is None or parent_id not in all_ids:
            roots.append(info)
        else:
            children_map[parent_id].append(eid)

    def build_node(node_info):
        node = {
            'id': node_info['id'],
            'code': node_info['code'],
            'name': node_info['name'],
            'level': node_info['level'],
            'amount': 0.0,
            'children': {}
        }
        for child_id in children_map.get(node_info['id'], []):
            child_node = build_node(nodes_info[child_id])
            node['children'][child_node['code']] = child_node
        return node

    return [build_node(info) for info in roots if info['parent_id'] is None or info['parent_id'] not in all_ids]

def _aggregate_tree(tree):
    """Рекурсивно вычисляет сумму узла как сумму его потомков."""
    for node in tree:
        if node['children']:
            _aggregate_tree(list(node['children'].values()))
            node['amount'] = round(sum(child['amount'] for child in node['children'].values()), 2)

def add_hierarchical_numbers(tree, prefix=''):
    """Присваивает узлам иерархический номер."""
    for idx, node in enumerate(tree, 1):
        node['hierarchical_number'] = prefix + str(idx) if prefix else str(idx)
        if node['children']:
            add_hierarchical_numbers(list(node['children'].values()), node['hierarchical_number'] + '.')

def assign_amounts(tree, top_element_ids, total_amount):
    """
    Распределяет сумму поровну между top_element_ids,
    затем рекурсивно распределяет долю каждого вниз по потомкам.
    """
    if not top_element_ids:
        return tree
    amount_per_top = total_amount / len(top_element_ids)

    def distribute_node(node, amount, is_target=False):
        if node['id'] in top_element_ids:
            if node['children']:
                child_amount = amount / len(node['children'])
                for child in node['children'].values():
                    distribute_node(child, child_amount, True)
            else:
                node['amount'] = round(float(amount), 2)
        else:
            if node['children']:
                child_amount = amount / len(node['children'])
                for child in node['children'].values():
                    distribute_node(child, child_amount, is_target)
            else:
                node['amount'] = round(float(amount), 2)

    for root in tree:
        distribute_node(root, amount_per_top, root['id'] in top_element_ids)

    _aggregate_tree(tree)
    add_hierarchical_numbers(tree)
    return tree

def calculate_distribution_new(element_ids, total_amount, version_date):
    """Основная функция расчёта распределения."""
    if not element_ids:
        return {}
    top_elements = filter_independent_elements(element_ids, version_date)
    if not top_elements:
        return {}
    tree = build_full_tree(top_elements, version_date)
    total = Decimal(str(total_amount))
    tree = assign_amounts(tree, top_elements, total)
    result = {}
    for root in tree:
        result[root['code']] = root
    return result

# ============================================================================
# API Endpoints
# ============================================================================

@app.route('/api/health', methods=['GET'])
def health_check():
    try:
        db.session.execute(db.text('SELECT 1'))
        redis_client.ping()
        return jsonify({'status': 'healthy', 'timestamp': datetime.utcnow().isoformat()}), 200
    except Exception as e:
        logger.error(f"Health check failed: {str(e)}")
        return jsonify({'status': 'unhealthy', 'error': str(e)}), 503

@app.route('/api/auth/login', methods=['POST'])
def login():
    user_id = request.json.get('user_id') or f'user-{uuid.uuid4().hex[:8]}'
    session_id = str(uuid.uuid4())
    access_token = create_access_token(identity=user_id)
    user_session = UserSession(
        session_id=session_id,
        user_id=user_id,
        jwt_token=access_token,
        expires_at=datetime.utcnow() + timedelta(days=30)
    )
    db.session.add(user_session)
    db.session.commit()
    return jsonify({
        'access_token': access_token,
        'session_id': session_id,
        'user_id': user_id
    }), 200

@app.route('/api/spp/structure', methods=['GET'])
def get_spp_structure():
    try:
        date_str = request.args.get('date')
        if date_str:
            try:
                target_date = datetime.strptime(date_str, '%Y-%m-%d')
            except ValueError:
                return jsonify({'success': False, 'error': 'Invalid date format. Use YYYY-MM-DD'}), 400

            query = db.text("""
                SELECT 
                    h.element_id AS id,
                    h.code,
                    h.name,
                    h.description,
                    h.parent_id,
                    h.status,
                    h.level
                FROM spp_history h
                WHERE h.valid_from <= :target_date
                  AND (h.valid_to IS NULL OR h.valid_to > :target_date)
                  AND h.status = 'ACTIVE'
                ORDER BY h.level, h.code
            """)
            rows = db.session.execute(query, {'target_date': target_date}).mappings().fetchall()
            nodes = [dict(row) for row in rows]
            structure = _build_tree_from_dicts(nodes)
        else:
            root_elements = SPPElement.query.filter_by(parent_id=None, status='ACTIVE').all()
            structure = [_build_tree(element) for element in root_elements]
        return jsonify({
            'success': True,
            'structure': structure,
            'timestamp': datetime.utcnow().isoformat()
        }), 200
    except Exception as e:
        logger.error(f"Error fetching SPP structure: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

def _build_tree_from_dicts(nodes):
    node_map = {}
    for node in nodes:
        node['children'] = []
        node_map[node['id']] = node
    roots = []
    for node in nodes:
        parent_id = node['parent_id']
        if parent_id is None:
            roots.append(node)
        else:
            parent = node_map.get(parent_id)
            if parent:
                parent['children'].append(node)
    for node in nodes:
        if not node['children']:
            del node['children']
    return roots

def _build_tree(element, include_inactive=False):
    tree = {
        'id': element.id,
        'code': element.code,
        'name': element.name,
        'description': element.description,
        'status': element.status,
        'level': element.level
    }
    query = SPPElement.query.filter_by(parent_id=element.id)
    if not include_inactive:
        query = query.filter_by(status='ACTIVE')
    children = query.all()
    if children:
        tree['children'] = [_build_tree(child, include_inactive) for child in children]
    return tree

@app.route('/api/spp/available-dates', methods=['GET'])
def get_available_dates():
    try:
        query = db.text("""
            SELECT DISTINCT DATE(valid_from) as version_date
            FROM spp_history
            ORDER BY version_date DESC
            LIMIT 30
        """)
        rows = db.session.execute(query).mappings().fetchall()
        dates = [row['version_date'].isoformat() for row in rows]
        return jsonify({'success': True, 'dates': dates}), 200
    except Exception as e:
        logger.error(f"Error fetching available dates: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/distribution/calculate', methods=['POST'])
@jwt_required()
def calculate_distribution():
    try:
        user_id = get_jwt_identity()
        data = request.json
        element_ids = data.get('element_ids', [])
        total_amount = float(data.get('total_amount', 0))
        version_date_str = data.get('version_date')

        if not element_ids or total_amount <= 0:
            return jsonify({'success': False, 'error': 'Invalid element_ids or total_amount'}), 400

        version_date = None
        if version_date_str:
            try:
                version_date = datetime.strptime(version_date_str, '%Y-%m-%d')
            except ValueError:
                return jsonify({'success': False, 'error': 'Invalid version_date format'}), 400
        else:
            version_date = datetime.utcnow()

        distribution_result = calculate_distribution_new(element_ids, total_amount, version_date)

        result_id = str(uuid.uuid4())
        redis_data = {
            'user_id': user_id,
            'timestamp': datetime.utcnow().isoformat(),
            'version_date': version_date.strftime('%Y-%m-%d'),
            'total_amount': total_amount,
            'element_ids': element_ids,
            'distribution': distribution_result
        }
        redis_client.setex(f'distribution:{result_id}', 86400, json.dumps(redis_data, default=str))

        return jsonify({
            'success': True,
            'result_id': result_id,
            'distribution': distribution_result,
            'total_amount': total_amount
        }), 200
    except Exception as e:
        logger.error(f"Error calculating distribution: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/distribution/save', methods=['POST'])
@jwt_required()
def save_distribution():
    try:
        user_id = get_jwt_identity()
        data = request.json
        result_id = data.get('result_id')
        session_id = data.get('session_id')
        if not result_id:
            return jsonify({'success': False, 'error': 'result_id required'}), 400

        redis_key = f'distribution:{result_id}'
        redis_data = redis_client.get(redis_key)
        if not redis_data:
            return jsonify({'success': False, 'error': 'Distribution result not found'}), 404
        redis_data = json.loads(redis_data)

        version_date = None
        if redis_data.get('version_date'):
            version_date = datetime.fromisoformat(redis_data['version_date'])

        distribution = DistributionResult(
            session_id=session_id or str(uuid.uuid4()),
            version_date=version_date or datetime.utcnow(),
            total_amount=redis_data['total_amount'],
            distribution_data=redis_data['distribution'],
            status='SAVED',
            meta_info={
                'user_id': user_id,
                'element_ids': redis_data['element_ids']
            }
        )
        db.session.add(distribution)
        db.session.commit()

        socketio.emit('distribution_saved', {
            'result_id': str(distribution.id),
            'session_id': distribution.session_id,
            'timestamp': datetime.utcnow().isoformat()
        }, room=session_id)

        return jsonify({
            'success': True,
            'id': distribution.id,
            'session_id': distribution.session_id
        }), 201
    except Exception as e:
        logger.error(f"Error saving distribution: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/distribution/saved-results', methods=['GET'])
@jwt_required()
def get_saved_results():
    try:
        session_id = request.args.get('session_id')
        if not session_id:
            return jsonify({'success': False, 'error': 'session_id required'}), 400
        results = DistributionResult.query.filter_by(
            session_id=session_id, status='SAVED'
        ).order_by(DistributionResult.created_at.desc()).all()
        results_data = [{
            'id': r.id,
            'total_amount': float(r.total_amount),
            'created_at': r.created_at.isoformat(),
            'version_date': r.version_date.isoformat() if r.version_date else None
        } for r in results]
        return jsonify({'success': True, 'results': results_data}), 200
    except Exception as e:
        logger.error(f"Error fetching saved results: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/distribution/<int:result_id>/load', methods=['GET'])
@jwt_required()
def load_distribution(result_id):
    try:
        result = DistributionResult.query.get(result_id)
        if not result:
            return jsonify({'success': False, 'error': 'Distribution result not found'}), 404
        return jsonify({
            'success': True,
            'id': result.id,
            'total_amount': float(result.total_amount),
            'distribution': result.distribution_data,
            'created_at': result.created_at.isoformat()
        }), 200
    except Exception as e:
        logger.error(f"Error loading distribution: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/distribution/<int:result_id>/export', methods=['GET'])
@jwt_required()
def export_distribution_excel(result_id):
    try:
        result = DistributionResult.query.get(result_id)
        if not result:
            return jsonify({'success': False, 'error': 'Distribution result not found'}), 404

        wb = Workbook()
        ws = wb.active
        ws.title = "Distribution"

        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 35

        headers = ['Номер', 'Код', 'Наименование', 'Сумма', 'Отделы']
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row = 2
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        def flatten_distribution(dist_data):
            nonlocal row
            for code, node in dist_data.items():
                hier_num = node.get('hierarchical_number', '')
                level = node.get('level', 1)
                departments_str = ''
                try:
                    dept_query = db.text("""
                        SELECT d.name
                        FROM departments d
                        JOIN spp_department_bindings sdb ON d.id = sdb.department_id
                        WHERE sdb.spp_element_id = :element_id
                          AND sdb.valid_from <= :version_date
                          AND (sdb.valid_to IS NULL OR sdb.valid_to > :version_date)
                    """)
                    dept_rows = db.session.execute(dept_query, {
                        'element_id': node['id'],
                        'version_date': result.version_date
                    }).mappings().fetchall()
                    departments_str = ', '.join([r['name'] for r in dept_rows])
                except Exception:
                    departments_str = ''

                ws.cell(row=row, column=1, value=hier_num)
                ws.cell(row=row, column=2, value=node['code'])
                ws.cell(row=row, column=3, value=node['name'])
                ws.cell(row=row, column=4, value=node['amount'])
                ws.cell(row=row, column=5, value=departments_str)

                for col in range(1, 6):
                    ws.cell(row=row, column=col).border = border
                ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=row, column=3).alignment = Alignment(indent=level - 1, horizontal="left", vertical="center")
                row += 1
                if 'children' in node:
                    flatten_distribution(node['children'])

        flatten_distribution(result.distribution_data)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        result.exported_at = datetime.utcnow()
        result.status = 'EXPORTED'
        db.session.commit()

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'distribution_{result_id}_{datetime.utcnow().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    except Exception as e:
        logger.error(f"Error exporting distribution: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

# ============================================================================
# WebSocket Events
# ============================================================================

@socketio.on('connect')
def handle_connect():
    logger.info(f"Client connected: {request.sid}")
    emit('connected', {'data': 'Connected to server'})

@socketio.on('join_session')
def on_join_session(data):
    session_id = data.get('session_id')
    if session_id:
        join_room(session_id)
        emit('joined_session', {'session_id': session_id, 'timestamp': datetime.utcnow().isoformat()})
        logger.info(f"Client {request.sid} joined session {session_id}")

@socketio.on('leave_session')
def on_leave_session(data):
    session_id = data.get('session_id')
    if session_id:
        leave_room(session_id)
        logger.info(f"Client {request.sid} left session {session_id}")

@socketio.on('disconnect')
def handle_disconnect():
    logger.info(f"Client disconnected: {request.sid}")

@app.errorhandler(404)
def not_found(e):
    return jsonify({'error': 'Not found'}), 404

@app.errorhandler(500)
def server_error(e):
    logger.error(f"Server error: {str(e)}")
    return jsonify({'error': 'Internal server error'}), 500

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    socketio.run(app, host='0.0.0.0', port=5000, debug=False)